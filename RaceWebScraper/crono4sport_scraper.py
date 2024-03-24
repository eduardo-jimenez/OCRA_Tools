from athlete_data import AthleteData
from athlete_data import fillExcelWorksheet
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook
from copy import copy
from scraper_common import copy_cells, AthleteSorting, CalculateAGAthletePoints
import requests
import time
import os


# This function scrapes an event at the given division for the given sex
def ScrapeRaceResults(driver, url: str, elite_points, raceTitle: str, category: str, sex: str) -> list:

    print("Scraping Race results from crono4sport.es for ", raceTitle, " - ", category, " - ", sex, ": ", url)

    category_index = 0
    match category:
        case "Elite":
            category_index = 0
        case "GGEE":
            category_index = 1
    sex_value = "General"
    match sex:
        case "Masc":
            sex_value = "GeneralMasc"
        case "Fem":
            sex_value = "GeneralFem"
    
    driver.get(url)
    time.sleep(0.25)

    # pick the category
    category_selector_item = driver.find_element(by=By.NAME, value="Carrera")
    #category_items = category_selector_item.find_elements(by=By.TAG_NAME, value="option")
    #selected_category_list_item = category_items[category_index]
    #selected_category_list_item.click()
    category_selector = Select(category_selector_item)
    category_selector.select_by_index(category_index)
    time.sleep(0.25)

    # and the sex
    selector_sex_item = driver.find_element(by=By.NAME, value="Categoria")
    selector_sex = Select(selector_sex_item)
    selector_sex.select_by_value(sex_value)
    time.sleep(0.25)

    # now we should iterate over all the links to athletes
    debugCounter = 0
    athletes = []

    # scroll down to the end of the table
    html = driver.find_element(By.TAG_NAME, 'html')
    table_athletes_item = driver.find_element(by=By.ID, value="inscritos")
    table_athletes_body_item = table_athletes_item.find_element(by=By.TAG_NAME, value="tbody")
    athletes_table_items = table_athletes_body_item.find_elements(by=By.TAG_NAME, value="tr")
    numAthletes = len(athletes_table_items)
    for i in range(5):
        table_head_item = table_athletes_item.find_element(by=By.TAG_NAME, value="thead")
        table_head_item.click()
        html.send_keys(Keys.END)
        time.sleep(0.5)

        table_athletes_item = driver.find_element(by=By.ID, value="inscritos")
        table_athletes_body_item = table_athletes_item.find_element(by=By.TAG_NAME, value="tbody")
        athletes_table_items = table_athletes_body_item.find_elements(by=By.TAG_NAME, value="tr")
        newNumAthletes = len(athletes_table_items)
        if numAthletes >= newNumAthletes:
            print("Found the end of the page after scrolling " + str(i + 1) + " times")
            break
        else:
            numAthletes = newNumAthletes

    table_athletes_item = driver.find_element(by=By.ID, value="inscritos")
    table_athletes_body_item = table_athletes_item.find_element(by=By.TAG_NAME, value="tbody")
    athletes_table_items = table_athletes_body_item.find_elements(by=By.TAG_NAME, value="tr")
    print("Analyzing " + str(len(athletes_table_items)) + " athlete rows")

    for athlete_root_item in athletes_table_items:
        athlete_parts = athlete_root_item.find_elements(by=By.TAG_NAME, value="td")

        # the first item is the position
        pos = athlete_parts[0].text
        # the second is the number and the third the name
        number = athlete_parts[1].text
        athleteName = athlete_parts[2].text
        athleteCat = athlete_parts[3].text
        # the seventh is the category position and the ninth the time
        categoryPos = athlete_parts[4].text
        clubName = athlete_parts[5].text
        timeStr = athlete_parts[7].text
        if (pos == 'DESC'):
            timeStr = 'Descalificado'

        if debugCounter % 10 == 0:
            print(pos + " " + number + " - " + athleteName + "( " + clubName + " ). Finished " + categoryPos + " in " + athleteCat + " with a time of " + timeStr)
        debugCounter += 1

        # fill the athlete data
        athlete = AthleteData()
        athlete.number = number
        athlete.name = athleteName
        athlete.club = clubName
        athlete.category = athleteCat
        athlete.posInCategory = categoryPos
        athlete.timeStr = timeStr
        athlete.computeTimeInMs()

        # apend the athlete to the list
        athletes.append(athlete)

    # after having all the athletes let's compute their points
    numDisqualifiedAthletes = 0
    if category_index == 0:
        # for elite athletes we use the elite_points array but first we have to sort them (removing the disqualified ones)
        non_disqualified_athletes = []
        for athlete in athletes:
            if athlete.timeSecs > 0:
                non_disqualified_athletes.append(athlete)
            else:
                numDisqualifiedAthletes += 1
        
        # now sorrt the non disqualified athletes
        non_disqualified_athletes.sort(key=AthleteSorting)

        # now assign points
        minLength = min(len(non_disqualified_athletes), len(elite_points))
        for i in range(minLength):
            non_disqualified_athletes[i].points = elite_points[i]

    else:
        # for GGEE we need to find the fastest time and the points are based on that time
        minTime = 3600.0 * 24.0 * 365.0     # one year seems a big enough time
        for athlete in athletes:
            if athlete.timeSecs > 0:
                if athlete.timeSecs < minTime:
                    minTime = athlete.timeSecs
            else:
                numDisqualifiedAthletes += 1

        # now let's calculate the points of each athlete
        for athlete in athletes:
            if athlete.timeSecs > 0:
                athlete.points = CalculateAGAthletePoints(athlete.timeSecs, minTime)
            else:
                athlete.points = 0

    print("Finished analizying athletes. A total of " + str(len(athletes)) + " athletes analyzed. " + str(numDisqualifiedAthletes) + " of them are disqualified and get no points")

    # return the list of athletes
    return athletes


def ScrapeCrono4SportFullRace(driver, url: str, elite_points, eventName: str, excelFilePath: str):

    print("Scraping results from crono4sports webpage")

    # Create an Excel file
    workbook = Workbook()
    origWorksheet = workbook.active
    

    # Copy the stats sheet from the reference sheet
    # reference_workbook = load_workbook('data\\RefRaceSheet.xlsx')
    # reference_sheet = reference_workbook.get_sheet_by_name("Stats")
    # worksheet.title = "Stats"
    # copy_cells(reference_sheet, worksheet)

    # Elite Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Masc')
    worksheet = workbook.create_sheet("Elite_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # Elite Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Fem')
    worksheet = workbook.create_sheet("Elite_Fem")
    fillExcelWorksheet(worksheet, athletes)

    # AG Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Masc')
    worksheet = workbook.create_sheet("GGEE_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # AG Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Fem')
    worksheet = workbook.create_sheet("GGEE_Fem")
    fillExcelWorksheet(worksheet, athletes)

    workbook.remove_sheet(origWorksheet)

    workbook.save(excelFilePath)
