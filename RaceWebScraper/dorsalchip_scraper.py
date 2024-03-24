from athlete_data import AthleteData
from athlete_data import fillExcelWorksheet
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook
from copy import copy
from scraper_common import copy_cells, AthleteSorting, CalculateAGAthletePoints
import requests
import time
import os


# This function scrapes an event at the given division for the given sex
def ScrapeRaceResults(driver: webdriver.Chrome, url: str, elite_points: list, raceTitle: str, category: str, sex: str) -> list:

    print("Scraping Race results from crono4sport.es for ", raceTitle, " - ", category, " - ", sex, ": ", url)

    category_str = category.upper()
    sex_str = "M"
    match sex:
        case "Masc":
            sex_str = "M"
        case "Fem":
            sex_str = "F"
    
    driver.get(url)
    time.sleep(0.25)

    # select the results
    results_elem = driver.find_element(by=By.ID, value='resultados')
    results_elem.click()
    time.sleep(0.25)
    meta_elem = driver.find_element(by=By.ID, value='META')
    meta_elem.click()
    time.sleep(0.25)

    # pick the category
    table_categories = driver.find_element(by=By.ID, value="META_modalidad")
    table_categories_links = table_categories.find_elements(by=By.TAG_NAME, value="a")
    selected_category_link = None
    for table_category_link in table_categories_links:
        text_upper = table_category_link.text.upper()
        if (category_str in text_upper) and (sex_str in text_upper):
            selected_category_link = table_category_link
            break
    # check we found the link
    if (selected_category_link == None):
        print("Error! Couldn't find the category: " + category_str + " " + sex_str)
        return None
    
    # click the link
    selected_category_link.click()
    time.sleep(0.25)

    # start scraping athlete results
    athletes = []
    next_button_disabled = False
    while not next_button_disabled:
        # scrape the athlete results from this page
        table_athletes_item = driver.find_element(by=By.ID, value="ResultadoJsonMeta")
        table_athletes_body_item = table_athletes_item.find_element(by=By.TAG_NAME, value="tbody")
        all_athletes_table_items = table_athletes_body_item.find_elements(by=By.TAG_NAME, value="tr")
        athletes_table_items = []
        for athlete_table_item in all_athletes_table_items:
            item_role = athlete_table_item.get_attribute("role")
            item_class = athlete_table_item.get_attribute("class")
            if item_class == "odd" or item_class == "even":
                athletes_table_items.append(athlete_table_item)
        print("Analyzing " + str(len(athletes_table_items)) + " athlete rows")

        for athlete_table_item in athletes_table_items:
            # get all the parts of the row
            athlete_parts = athlete_table_item.find_elements(by=By.TAG_NAME, value="td")

            # the first item is the overall position
            pos = athlete_parts[0].text

            # the second item is the number
            number = athlete_parts[1].text

            # the third item is the name and the club
            name_and_club = athlete_parts[2].text
            name_newline_index = name_and_club.find('\n')
            athleteName = name_and_club[:name_newline_index]
            clubName = name_and_club[name_newline_index + 1:]
            if clubName[0] == '*':
                clubName = clubName[1:]

            # the fourth item has the time
            time_item_str = athlete_parts[3].text
            total_str_index = time_item_str.find("Total:")
            timeStr = time_item_str[total_str_index + 6:total_str_index + 14]

            # the fifth item has the category and the category position
            cat_and_pos = athlete_parts[4].text
            tanda_str_index = cat_and_pos.find("TANDA")
            categoryPos = cat_and_pos[0:tanda_str_index - 2]
            athleteCat = cat_and_pos[tanda_str_index + 7:]
            if athleteCat[0] == ' ':
                athleteCat = athleteCat[1:]

            print(pos + " " + number + " - " + athleteName + "( " + clubName + " ). Finished " + categoryPos + " in " + athleteCat + " with a time of " + timeStr)

            # fill the athlete data
            athlete = AthleteData()
            athlete.number = number
            athlete.name = athleteName
            athlete.club = clubName
            athlete.category = athleteCat
            athlete.posInCategory = categoryPos
            athlete.timeStr = timeStr
            athlete.computeTimeInMs()

            # apend the athlete to the list
            athletes.append(athlete)

        # if the next page link is active, press it
        next_page_link = driver.find_element(by=By.ID, value="ResultadoJsonMeta_next")
        next_page_link_class = next_page_link.get_attribute("class")
        next_button_disabled = "disabled" in next_page_link_class
        if not next_button_disabled:
            next_page_link.click()
            time.sleep(0.25)

    # after having all the athletes let's compute their points
    if category_str == "ELITE":
        # sort the non disqualified athletes
        athletes.sort(key=AthleteSorting)

        # now assign points
        minLength = min(len(athletes), len(elite_points))
        for i in range(minLength):
            athletes[i].points = elite_points[i]

    else:
        # for GGEE we need to find the fastest time and the points are based on that time
        minTime = 3600.0 * 24.0 * 365.0     # one year seems a big enough time
        for athlete in athletes:
            if athlete.timeSecs > 0:
                if athlete.timeSecs < minTime:
                    minTime = athlete.timeSecs

        # now let's calculate the points of each athlete
        for athlete in athletes:
            if athlete.timeSecs > 0:
                athlete.points = CalculateAGAthletePoints(athlete.timeSecs, minTime)
            else:
                athlete.points = 0

    print("Finished analizying athletes. A total of " + str(len(athletes)) + " athletes analyzed.")

    # return the list of athletes
    return athletes


def ScrapeDorsalChipFullRace(driver: webdriver.Chrome, url: str, elite_points: list, eventName: str, excelFilePath: str):

    print("Scraping results from dorsalchip webpage")

    # Create an Excel file
    workbook = Workbook()
    origWorksheet = workbook.active
    

    # Copy the stats sheet from the reference sheet
    # reference_workbook = load_workbook('data\\RefRaceSheet.xlsx')
    # reference_sheet = reference_workbook.get_sheet_by_name("Stats")
    # worksheet.title = "Stats"
    # copy_cells(reference_sheet, worksheet)

    # Elite Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Masc')
    worksheet = workbook.create_sheet("Elite_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # Elite Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Fem')
    worksheet = workbook.create_sheet("Elite_Fem")
    fillExcelWorksheet(worksheet, athletes)

    # AG Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Masc')
    worksheet = workbook.create_sheet("GGEE_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # AG Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Fem')
    worksheet = workbook.create_sheet("GGEE_Fem")
    fillExcelWorksheet(worksheet, athletes)

    workbook.remove_sheet(origWorksheet)

    workbook.save(excelFilePath)
