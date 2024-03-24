import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from athlete_data import AthleteData


class AthleteRaceInfo:
    def __init__(self):
        self.race_name = ""
        self.timeInRace = ""
        self.position = ""
        self.pointsInRace = 0

class LeagueAthlete:
    def __init__(self):
        self.number = -1
        self.name = ""
        self.date_of_birth = None
        self.age_group = ""
        self.club = ""
        self.category = ""
        self.races_considered = 0
        self.points = 0
        self.races: List[AthleteRaceInfo] = []

    def calculate_age_group(self):
        # calculate how many years old the athlete will be on 2024
        year_of_birth = self.date_of_birth.year
        years_old = 2024 - year_of_birth
        if years_old < 20:
            self.age_group = 'U20'
        elif years_old < 25:
            self.age_group = '20-24'
        elif years_old < 30:
            self.age_group = '25-29'
        elif years_old < 35:
            self.age_group = '30-34'
        elif years_old < 40:
            self.age_group = '35-39'
        elif years_old < 45:
            self.age_group = '40-44'
        elif years_old < 50:
            self.age_group = '45-49'
        elif years_old < 55:
            self.age_group = '50-54'
        elif years_old < 60:
            self.age_group = '55-59'
        else:
            self.age_group = '60+'


def sort_league_athlete(athlete: LeagueAthlete):
    return athlete.points

def sort_race_info(race_info: AthleteRaceInfo):
    return race_info.pointsInRace


# Loads all OCRA athletes from the right excel file
def load_all_OCRA_athletes(path: str):

    athletes: List[LeagueAthlete] = []

    # load the excel file
    workbook = load_workbook(path)
    sheet: Worksheet = workbook.active

    # the first row is the header, the rest are the athletes
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):

        # ignore clubs
        category_str = row[5].lower()
        if not ('atleta' in category_str):
            continue

        # first column is the number
        athlete_number = row[0]
        # second column is the name
        athlete_name = row[1]
        # third column is the club
        athlete_club = row[2]
        # fourth is the date of birth
        athlete_date_of_birth = row[3]
        # fifth category is the sex and the seventh whether they are elite or not
        athlete_sex = row[4]
        is_elite = (row[6] != None and row[6] != '')
        if (is_elite):
            athlete_category = 'Elite'
        else:
            athlete_category = 'GGEE'
        if (athlete_sex == 'Mujer'):
            athlete_category = athlete_category + '_Fem'
        else:
            athlete_category = athlete_category + '_Masc'

        # ensure all parameters are valid
        if (athlete_number == None or athlete_number == "" or
            athlete_date_of_birth == None or athlete_date_of_birth == ""):
            continue

        # create an athlete
        athlete = LeagueAthlete()
        athlete.club = athlete_club
        athlete.name = athlete_name
        athlete.number = int(athlete_number)
        athlete.category = athlete_category
        if athlete_date_of_birth is str:
            athlete.date_of_birth = datetime.datetime.strptime(athlete_date_of_birth, '%y/%m/%d')
        else:
            athlete.date_of_birth = athlete_date_of_birth
        athlete.calculate_age_group()

        # add it to the list
        athletes.append(athlete)

        #print(f'New Athlete: {athlete.name} ({athlete.number}) - {athlete.date_of_birth}')

    print(f'A total of {len(athletes)} athletes have been read from {path}')

    return athletes


# Analyzes an excel sheet with the results of a race for a cateogry and returns the list of athletes analyzed
def analyze_race_category_sheet(sheet: Worksheet):
    category_str = sheet.title
    print('Analyzing sheet ' + sheet.title)

    category_athletes: List[AthleteData] = []
    # iterate over the rows of the sheet, ignoring the first one which is the title
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if (row[0] == None):
            continue
        
        athlete = AthleteData()
        athlete.number = row[0]
        athlete.name = row[1]
        athlete.club = row[2]
        athlete.category = category_str
        athlete.posInCategory = row[4]
        athlete.timeStr = str(row[5])
        if row[5] is datetime.timedelta:
            athlete.timeSecs = row[5].total_seconds()
        else:
            athlete.timeSecs = 0.0
        athlete.points = row[6]
        category_athletes.append(athlete)
    
    return category_athletes


# Analyzes an excel file with the results of a race and returns the list of athletes analyzed
def analyze_race_excel(file_path: str):
    print('Opening race excel at ' + file_path)

    race_athletes = []

    # load the excel file
    workbook = load_workbook(file_path)

    # go analyzing each of the 4 sheets it should have with names 'Elite_Masc', 'Elite_Fem', 'GGEE_Masc', 'GGEE_Fem'
    race_athletes = []
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        sheet = workbook.get_sheet_by_name(sheet_name)
        category_athletes = analyze_race_category_sheet(sheet)
        race_athletes.extend(category_athletes)

    print(f'A total of {len(race_athletes)} athletes have been extracted from the file {file_path}')

    return race_athletes


# adds the athlete results from a race (in race_athletes) to to the athletes list
def add_race_resutls_athletes(race_athletes: List[AthleteData], athletes: List[LeagueAthlete], race_name: str):
    
    # iterate over all the athletes in the race
    for athlete_info in race_athletes:

        # compose the race info for this athlete
        race_info = AthleteRaceInfo()
        race_info.race_name = race_name
        race_info.timeInRace = athlete_info.timeStr
        race_info.position = athlete_info.posInCategory
        race_info.pointsInRace = athlete_info.points

        # try to find the athlete in the athletes list
        athlete_number = int(athlete_info.number)
        athlete_data = next((x for x in athletes if x.number == athlete_number), None)

        # we're reading the list of valid athletes before, so if we can't find him he's not part of OCRA España
        if athlete_data == None:
            continue

        # if the category is not the right one we skip the athlete too (if it's an elite competing in age group we are not going to consider its results)
        if athlete_data.category != athlete_info.category:
            continue

        if athlete_data.club == "":
            # add some data we didn't get from the original excel file
            athlete_data.club = athlete_info.club

        # append the race to the athlete data
        athlete_data.races.append(race_info)
        

# Calculates the points each athlete
def calculate_points_all_athletes(athletes: List[LeagueAthlete], max_races_to_consider: int):
    for athlete in athletes:
        # reset the points
        athlete.points = 0
        
        if (len(athlete.races) <= max_races_for_points):
            # add the points in all races
            athlete.races_considered = len(athlete.races)
            for race_info in athlete.races:
                athlete.points += race_info.pointsInRace
        else:
            # we have to pick the best races
            athlete.races.sort(key=sort_race_info, reverse=True)
            athlete.races_considered = max_races_for_points
            for i in range(0, max_races_for_points):
                athlete.points += athlete.races[i].pointsInRace


# analyzes the results of all the excel files in the given folder 
def analyze_all_races_in_folder(folder: str, athletes: List[LeagueAthlete], max_races_to_consider: int):
    print(f'Analyzing all files in folder {folder}')

    # iterate over all the files in the given folder
    races = []
    for filename in os.listdir(folder):
        # compose the final path
        filepath = os.path.join(folder, filename)
        filename_extension = os.path.splitext(filename)[1]

        # we only analyze Excel files
        if (filename_extension == '.xlsx'):
            filename_without_ext = os.path.splitext(filename)[0]
            race_name = filename_without_ext.replace("-", " ").replace("_", " ")
            races.append(race_name)

            # analyze it
            race_athletes = analyze_race_excel(filepath)

            # add the race results to the overall athletes array
            add_race_resutls_athletes(race_athletes, athletes, race_name)

    # calculate the points of all athletes
    calculate_points_all_athletes(athletes, max_races_to_consider)

    return races


# writes the header of the sheet for a league
def write_header_for_sheet(sheet: Worksheet, race_names: List[str], row_offset: int, isElite: bool):
    # write the headers
    sheet.cell(row_offset + 0, 1).value = "Info Atleta"
    sheet.cell(row_offset + 1, 1).value = "Posición"
    sheet.cell(row_offset + 1, 2).value = "Puntos"
    sheet.cell(row_offset + 1, 3).value = "Carreras"
    sheet.cell(row_offset + 1, 4).value = "Dorsal"
    sheet.cell(row_offset + 1, 5).value = "Nombre"
    sheet.cell(row_offset + 1, 6).value = "Club"
    if isElite:
        col_offset = 7
    else:
        col_offset = 9
        sheet.cell(row_offset + 1, 7).value = "Categoría"
        sheet.cell(row_offset + 1, 8).value = "Posición"

    for i in range(0, len(race_names)):
        col_index = col_offset + 2 * i
        sheet.cell(row_offset + 0, col_index).value = race_names[i]
        if isElite:
            sheet.cell(row_offset + 1, col_index + 0).value = "Posición"
        else:
            sheet.cell(row_offset + 1, col_index + 0).value = "Tiempo"
        sheet.cell(row_offset + 1, col_index + 1).value = "Puntos"
        sheet.merge_cells(start_row=row_offset, start_column=col_index , end_row=row_offset, end_column=col_index + 1)

    if (isElite):
        sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=6)
    else:
        sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=8)
    
    for i in range(1, col_offset + 2 * len(race_names)):
        race_col = i - col_offset
        if race_col == -1 or (race_col > 0 and race_col % 2 == 1):
            sheet.cell(row_offset + 0, i).border = Border(top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN))
            sheet.cell(row_offset + 1, i).border = Border(bottom=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN))
        else:
            sheet.cell(row_offset + 0, i).border = Border(top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
            sheet.cell(row_offset + 1, i).border = Border(bottom=Side(border_style=BORDER_THIN))

    # set all the cells as bold
    bold_font = Font(bold=True)
    row = sheet[row_offset + 0]
    for cell in row:
        cell.font = bold_font
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')
    row = sheet[row_offset + 1]
    for cell in row:
        cell.font = bold_font
        cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')


# writes the given athlete at the given row in the given sheet with the info we have
def write_athlete_row(sheet: Worksheet, row: int, pos: int, athlete: LeagueAthlete, race_names: List[str], isElite: bool, age_group_pos: {}):
    
    # write the info of the athlete
    sheet.cell(row, 1).value = pos
    sheet.cell(row, 2).value = athlete.points
    sheet.cell(row, 2).number_format = '#.#'
    sheet.cell(row, 3).value = athlete.races_considered
    sheet.cell(row, 4).value = athlete.number
    sheet.cell(row, 5).value = athlete.name
    sheet.cell(row, 6).value = athlete.club

    if isElite:
        col_offset = 7
        sheet.cell(row, 6).border = Border(right=Side(border_style=BORDER_THIN))
    else:
        col_offset = 9
        sheet.cell(row, 7).value = athlete.age_group
        if athlete.points > 0:
            sheet.cell(row, 8).value = age_group_pos[athlete.age_group]
        sheet.cell(row, 8).border = Border(right=Side(border_style=BORDER_THIN))

    # now write the results in the races
    for i in range(len(race_names)):
        race_name = race_names[i]

        # try to find the results for this race
        race_results = next((x for x in athlete.races if x.race_name == race_name), None)
        if (race_results != None):
            if isElite:
                sheet.cell(row, col_offset + 2 * i + 0).value = race_results.position
            else:
                sheet.cell(row, col_offset + 2 * i + 0).value = race_results.timeInRace
            sheet.cell(row, col_offset + 2 * i + 1).value = race_results.pointsInRace
            sheet.cell(row, col_offset + 2 * i + 1).number_format = '#.#'
        
        # in any case format the cell
        sheet.cell(row, col_offset + 2 * i + 1).border = Border(right=Side(border_style=BORDER_THIN))


# fill all the athletes of a given category in a worksheet
def fill_league_sheet(sheet: Worksheet, athletes: List[LeagueAthlete], race_names: List[str]):

    # extract only the athletes of the right category
    category_str = sheet.title
    category_athletes = [x for x in athletes if x.category == category_str]

    # sort them by the points they have
    category_athletes.sort(key=sort_league_athlete, reverse=True)

    # write the header first
    sheet.column_dimensions['E'].width = 40
    sheet.column_dimensions['F'].width = 40
    isElite = 'Elite' in category_str
    write_header_for_sheet(sheet, race_names, 1, isElite)

    # in age group we're going to keep track of the current position in each age group
    age_group_pos = {}
    age_group_pos['U20'] = 0
    age_group_pos['20-24'] = 0
    age_group_pos['25-29'] = 0
    age_group_pos['30-34'] = 0
    age_group_pos['35-39'] = 0
    age_group_pos['40-44'] = 0
    age_group_pos['45-49'] = 0
    age_group_pos['50-54'] = 0
    age_group_pos['55-59'] = 0
    age_group_pos['60+'] = 0

    # now iterate over all the athletes
    for i in range(0, len(category_athletes)):
        athlete: LeagueAthlete = category_athletes[i]

        # update the athlete age group
        if not isElite:
            age_group_pos[athlete.age_group] += 1

        # add a row for this athlete
        write_athlete_row(sheet, 3 + i, i + 1, athlete, race_names, isElite, age_group_pos)


# generates the final excel file with all athletes in the league
def generate_excel_league(athletes: List[LeagueAthlete], race_names: List[str], path: str):

    # create or open the excel file 
    workbook = Workbook()
    origWorksheet = workbook.active

    # create a sheet for each cathegory (Elite_Masc, Elite_Fem, GGEE_Masc, GGEE_Fem)
    
    worksheet = workbook.create_sheet("Elite_Masc")
    fill_league_sheet(worksheet, athletes, race_names)

    worksheet = workbook.create_sheet("Elite_Fem")
    fill_league_sheet(worksheet, athletes, race_names)

    worksheet = workbook.create_sheet("GGEE_Masc")
    fill_league_sheet(worksheet, athletes, race_names)

    worksheet = workbook.create_sheet("GGEE_Fem")
    fill_league_sheet(worksheet, athletes, race_names)

    # remove the original worksheet which we haven't used
    workbook.remove_sheet(origWorksheet)

    # save the excel to file
    workbook.save(path)



# set the folder name
league_name = 'LigaOCRA'
#league_name = 'OCRSeries'
analyzeLigaOCRA = (league_name == 'LigaOCRA')
if analyzeLigaOCRA:
    files_folder = 'data\\LigaOCRA'
    max_races_for_points = 8
else:
    files_folder = 'data\\OCRSeries'
    max_races_for_points = 6

# create a full path for the folder
currFolder = os.getcwd()
path = os.path.join(currFolder, files_folder)

# load all OCRA athletes
all_athletes_path = os.path.join(currFolder, 'data\\atletas_OCRA.xlsx')
athletes: List[AthleteData] = load_all_OCRA_athletes(all_athletes_path)

# analyze all the files inside
race_names = analyze_all_races_in_folder(path, athletes, max_races_for_points)

# compose the final path of the final file
league_filepath = files_folder + '.xlsx'

# finally generate the final excel
generate_excel_league(athletes, race_names, league_filepath)
