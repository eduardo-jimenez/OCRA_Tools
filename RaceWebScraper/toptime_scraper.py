from athlete_data import AthleteData
from athlete_data import fillExcelWorksheet
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook
from copy import copy
from scraper_common import copy_cells, AthleteSorting, CalculateAGAthletePoints
import requests
import time
import os


# This function scrapes an event at the given division for the given sex
def ScrapeRaceResults(driver: webdriver.Chrome, url: str, elite_points: list, raceTitle: str, category: str, sex: str) -> list:

    print("Scraping Race results from crono4sport.es for ", raceTitle, " - ", category, " - ", sex, ": ", url)

    is_elite = 'elite' in category.lower()
    sex_elem_id = "mnuHF"
    match sex:
        case "Masc":
            sex_elem_id = "mnuHom"
        case "Fem":
            sex_elem_id = "mnuFem"
    
    driver.get(url)
    time.sleep(0.5)

    iframe_item = driver.find_element(by=By.ID, value='FGL')
    new_url = iframe_item.get_attribute('src')
    driver.get(new_url)
    time.sleep(1.0)

    # select the sex
    sex_elem = driver.find_element(by=By.ID, value='mnuSx')
    sex_elem.click()
    time.sleep(0.25)
    meta_elem = sex_elem.find_element(by=By.ID, value=sex_elem_id)
    meta_elem.click()
    time.sleep(0.25)

    # start scraping athlete results
    athletes = []

    # scrape the athlete results from this page
    table_athletes_item = driver.find_element(by=By.ID, value="tabres")
    table_athletes_body_item = table_athletes_item.find_element(by=By.TAG_NAME, value="tbody")
    athletes_table_items = table_athletes_body_item.find_elements(by=By.TAG_NAME, value="tr")
    print("Analyzing " + str(len(athletes_table_items)) + " athlete rows")

    for athlete_table_item in athletes_table_items:
        # get all the parts of the row
        athlete_parts = athlete_table_item.find_elements(by=By.TAG_NAME, value="td")

        # there should be at least 12 parts (usually 14)
        if len(athlete_parts) < 12:
            continue

        # first let's check this is an OCRA competitor
        is_ocra_str = athlete_parts[7].text
        if (is_ocra_str != 'SI'):
            continue
        number = athlete_parts[8].text
        if (number == ""):
            continue

        # now check if it elite/GGEE
        athleteCat = athlete_parts[5].text
        if is_elite and not ('Elite' in athleteCat):
            continue
        elif not is_elite and ('Elite' in athleteCat or 'Pop' in athleteCat):
            continue

        # the first item is the overall position
        pos = athlete_parts[0].text

        # the third item is the name and a flag for the nationality
        athlete_name = athlete_parts[2].text

        # the fourth is the club
        club_name = athlete_parts[3].text

        # the seventh is the position in the category
        categoryPos = athlete_parts[6].text

        # the eleventh item has the time
        timeStr = athlete_parts[10].text

        print(pos + " " + number + " - " + athlete_name + "( " + club_name + " ). Finished " + categoryPos + " in " + athleteCat + " with a time of " + timeStr)

        # fill the athlete data
        athlete = AthleteData()
        athlete.number = number
        athlete.name = athlete_name
        athlete.club = club_name
        athlete.category = athleteCat
        athlete.posInCategory = categoryPos
        athlete.timeStr = timeStr
        athlete.computeTimeInMs()

        # apend the athlete to the list
        athletes.append(athlete)

    # after having all the athletes let's compute their points
    numDisqualifiedAthletes = 0
    if is_elite:
        # for elite athletes we use the elite_points array but first we have to sort them (removing the disqualified ones)
        non_disqualified_athletes = []
        for athlete in athletes:
            if athlete.timeSecs > 0:
                non_disqualified_athletes.append(athlete)
            else:
                numDisqualifiedAthletes += 1
        
        # now sorrt the non disqualified athletes
        non_disqualified_athletes.sort(key=AthleteSorting)

        # now assign points
        minLength = min(len(non_disqualified_athletes), len(elite_points))
        for i in range(minLength):
            non_disqualified_athletes[i].points = elite_points[i]

    else:
        # for GGEE we need to find the fastest time and the points are based on that time
        minTime = 3600.0 * 24.0 * 365.0     # one year seems a big enough time
        for athlete in athletes:
            if athlete.timeSecs > 0:
                if athlete.timeSecs < minTime:
                    minTime = athlete.timeSecs
            else:
                numDisqualifiedAthletes += 1

        # now let's calculate the points of each athlete
        for athlete in athletes:
            if athlete.timeSecs > 0:
                athlete.points = CalculateAGAthletePoints(athlete.timeSecs, minTime)
            else:
                athlete.points = 0

    print("Finished analizying athletes. A total of " + str(len(athletes)) + " athletes analyzed.")

    # return the list of athletes
    return athletes


def ScrapeToptimeFullRace(driver: webdriver.Chrome, url: str, elite_points: list, eventName: str, excelFilePath: str):

    print("Scraping results from dorsalchip webpage")

    # Create an Excel file
    workbook = Workbook()
    origWorksheet = workbook.active
    

    # Copy the stats sheet from the reference sheet
    # reference_workbook = load_workbook('data\\RefRaceSheet.xlsx')
    # reference_sheet = reference_workbook.get_sheet_by_name("Stats")
    # worksheet.title = "Stats"
    # copy_cells(reference_sheet, worksheet)

    # Elite Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Masc')
    worksheet = workbook.create_sheet("Elite_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # Elite Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'Elite', 'Fem')
    worksheet = workbook.create_sheet("Elite_Fem")
    fillExcelWorksheet(worksheet, athletes)

    # AG Men
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Masc')
    worksheet = workbook.create_sheet("GGEE_Masc")
    fillExcelWorksheet(worksheet, athletes)

    # AG Women
    athletes = ScrapeRaceResults(driver, url, elite_points, eventName, 'GGEE', 'Fem')
    worksheet = workbook.create_sheet("GGEE_Fem")
    fillExcelWorksheet(worksheet, athletes)

    workbook.remove_sheet(origWorksheet)

    workbook.save(excelFilePath)
