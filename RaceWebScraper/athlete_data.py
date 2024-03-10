import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from datetime import timedelta


# Converts a string with format '##:##:##' to the total seconds
def timeStrToMSeconds(str: str) -> int:
    if (len(str) <= 1):
        return 0
    
    parts = re.findall(r'\d+', str)
    hours = int(parts[0])
    minutes = int(parts[1])
    seconds = int(parts[2])
    centsSec = 0
    if (len(parts) >= 4):
        centsSec = int(parts[3])
    runningTime = hours * 3600000 + minutes * 60000 + seconds * 1000 + 10 * centsSec

    return runningTime


# Converts an integer to a time with format '##:##:##' considering the time as seconds
def timeMSecondsToStr(totalMSeconds: int) -> str:
    msecs = totalMSeconds % 1000
    totalSeconds = totalMSeconds / 1000;
    seconds = totalSeconds
    hours = int(seconds // 3600)
    seconds -= hours * 3600
    minutes = int(seconds // 60)
    seconds -= minutes * 60
    seconds = int(seconds)
    timeStr = str(hours).zfill(2) + ":" + str(minutes).zfill(2) + ":" + str(seconds).zfill(2) + "." + str(msecs).zfill(3)

    return timeStr


class AthleteData:
    def __init__(self):
        self.number = -1
        self.name = ""
        self.club = ""
        self.category = ""
        self.posInCategory = ""
        self.timeStr = ""
        self.timeSecs = 0.0
        self.points = 0


    def getCSVHeader():
        str = f'Pos;Dorsal;Nombre;Club;Categoría;Pos en Categoría;Tiempo;Puntos'
        return str

    def getCSVLine(self):
        str = f'\"{self.number}\";\"{self.name}\";\"{self.club}\";\"{self.category}\";{self.posInCategory};{self.timeStr};{self.points}'
        return str

    def computeTimeInMs(self):
        try:
            timeInMs = timeStrToMSeconds(self.timeStr)
        except:
            timeInMs = 0
        self.timeSecs = timeInMs / 1000.0

    def writeHeaderInExcelWorksheet(sheet: Worksheet, row: int):
        # write the headers
        sheet.cell(row, 1).value = "Dorsal"
        sheet.cell(row, 2).value = "Nombre"
        sheet.cell(row, 3).value = "Club"
        sheet.cell(row, 4).value = "Categoría"
        sheet.cell(row, 5).value = "Pos en Categoría"
        sheet.cell(row, 6).value = "Tiempo"
        sheet.cell(row, 7).value = "Puntos"

        # set all the cells as bold
        bold_font = Font(bold=True)
        row = sheet[row]
        for cell in row:
            cell.font = bold_font

    def writeAtheleteInfoInWorksheet(self, sheet: Worksheet, row: int):
        sheet.cell(row, 1).value = self.number
        sheet.cell(row, 2).value = self.name
        sheet.cell(row, 3).value = self.club
        sheet.cell(row, 4).value = self.category
        sheet.cell(row, 5).value = self.posInCategory
        if (self.timeSecs > 0):
            sheet.cell(row, 6).value = self.timeSecs / (24.0 * 3600.0)  # for some reason the time needs to be stored in 'days'
        else:
            sheet.cell(row, 6).value = self.timeStr
        sheet.cell(row, 7).value = self.points
        sheet.cell(row, 6).number_format = "[HH]:MM:SS.000"


# Writes a file with the info of all the athletes in the given list
def writeAtheletesToCSV(athletes: list[AthleteData], filePath: str):
    f = open(file=filePath, mode="w", encoding="utf-8")
    f.write(AthleteData.getCSVHeader() + '\n')
    for athlete in athletes:
        f.write(athlete.getCSVLine() + '\n')
    f.flush()
    f.close()
    print(f'Finishing writing {len(athletes)} athletes to file {filePath}')


def fillExcelWorksheet(sheet: Worksheet, athletes:list[AthleteData]):
    # write the header first
    AthleteData.writeHeaderInExcelWorksheet(sheet, 1)

    # now write all the athlete data
    row = 2
    for athlete in athletes:
        athlete.writeAtheleteInfoInWorksheet(sheet, row)
        row += 1

    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['F'].width = 15
